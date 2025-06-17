# This files contains your custom actions which can be used to run
# custom Python code.
#
# See this guide on how to implement these action:
# https://rasa.com/docs/rasa/custom-actions


# This is a simple example for a custom action which utters "Hello World!"

from typing import Any, Text, Dict, List

from rasa_sdk import Action, Tracker
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import UserUtteranceReverted
from datetime import datetime
import random
import os
import openpyxl


class ActionCheckAvailability(Action):
     """Check availability for requested dates"""

     def name(self) -> Text:
         return "action_check_availability"

     def run(self, dispatcher: CollectingDispatcher,
             tracker: Tracker,
             domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        # In a real implementation, you would check a database here
         dispatcher.utter_message(text="¡Si, hay disponibilidad para las fechas solicitadas1!")

         return []

class ActionProvidePricing(Action):
    """Provide detailed pricing information"""
    
    def name(self) -> Text:
        return "action_provide_pricing"
        
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        pricing_info = """
        Ecos de la Montaña tiene 3 modalidades de alojamiento:
        - Apartamento para 10 personas: $700.000 por noche
        - Habitación para 3 personas: $200.000 por noche
        - Micro-apartamento para 2 personas: $250.000 por noche
        """
        dispatcher.utter_message(text=pricing_info)
        
        return []

class ActionHandleBooking(Action):
    """Handle booking process"""
    
    def name(self) -> Text:
        return "action_handle_booking"
        
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        # Get slots or entities from tracker
        # date = tracker.get_slot("date")
        # accommodation_type = tracker.get_slot("accommodation_type")
        
        dispatcher.utter_message(text="¡Reserva recibida! Pronto te contactaremos para confirmar los detalles.")
        
        return []

class ActionProvideLocation(Action):
    """Provide location details"""
    
    def name(self) -> Text:
        return "action_provide_location"
        
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        location_info = """
        Estamos ubicados en:
        Km 24 vía Rionegro-Playon, Santander.
        Dentro del balneario El Portal (3 km por vía destapada desde la entrada).
        """
        dispatcher.utter_message(text=location_info)
        
        return []

class ActionHandleModification(Action):
    """Handle reservation modifications"""
    
    def name(self) -> Text:
        return "action_handle_modification"
        
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        dispatcher.utter_message(text="Hemos recibido tu solicitud de modificación. Te contactaremos pronto.")
        
        return []

class ActionHandleCancellation(Action):
    """Handle reservation cancellations"""
    
    def name(self) -> Text:
        return "action_handle_cancellation"
        
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        dispatcher.utter_message(text="Lamentamos que tengas que cancelar. Hemos registrado tu solicitud.")
        
        return []

class ActionProvidePaymentInfo(Action):
    """Provide payment information"""
    
    def name(self) -> Text:
        return "action_provide_payment_info"
        
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        payment_info = """
        Por favor realiza el pago a:
        Banco: Bancolombia
        Cuenta de ahorros: 0000000
        Nombre: Ecos de la Montaña
        """
        dispatcher.utter_message(text=payment_info)
        
        return []

class ActionListAmenities(Action):
    """List all available amenities"""
    
    def name(self) -> Text:
        return "action_list_amenities"
        
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        amenities = """
        Nuestros servicios incluyen:
        - Hidropiscina con tratamiento salino
        - Sauna
        - Zona de recreación
        - Kiosco con cocina
        - Wifi
        - Linos para uso en instalaciones
        - Acceso a pozos naturales de la quebrada Samacá
        """
        dispatcher.utter_message(text=amenities)
        
        return []

class ActionProvideCheckInOutInfo(Action):
    """Provide check-in/out information"""
    
    def name(self) -> Text:
        return "action_provide_check_in_out_info"
        
    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        check_info = """
        Horarios:
        - Check-in: 14:00 (2 PM)
        - Check-out: 12:00 (mediodía)
        """
        dispatcher.utter_message(text=check_info)
        
        return []   

class ActionDefaultFallback(Action):
    def name(self):
        return "action_default_fallback"

    async def run(self, dispatcher: CollectingDispatcher, tracker, domain):
        user_message = tracker.latest_message.get('text')
        # Ruta del archivo Excel
        file_path = "preguntas_no_entendidas.xlsx"

        # Si el archivo no existe, se crea con encabezado
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Preguntas"
            sheet['A1'] = "Pregunta del Usuario"
            workbook.save(file_path)

        # Abre y escribe la pregunta
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=user_message)
        workbook.save(file_path)

        dispatcher.utter_message(text="No entendí bien eso. ¿Podrías reformularlo?")
        return [UserUtteranceReverted()]